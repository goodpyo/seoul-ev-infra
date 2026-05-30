import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# 원본 파일에서 은평구 직접 확인
wb = openpyxl.load_workbook('전기차 충전소 설치현황_20260512.xlsx')
ws = wb.active

print("=== 충전소 파일에서 은평구 원본 데이터 ===")
ep_rows = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[2] and '은평구' in str(row[2]):
        ep_rows.append(row)

print(f"은평구 충전소 개소: {len(ep_rows)}곳")
total_fast = sum(int(str(r[3]).strip()) for r in ep_rows if r[3] and str(r[3]).strip().isdigit())
total_slow = sum(int(str(r[4]).strip()) for r in ep_rows if r[4] and str(r[4]).strip().isdigit())
print(f"급속: {total_fast}기 / 완속: {total_slow}기 / 합계: {total_fast+total_slow}기")
print()
for r in ep_rows:
    print(f"  {r[1]} | {r[2]} | 급속:{r[3]} 완속:{r[4]}")

print()
print("=== 파일에 포함된 전체 자치구 목록 ===")
import re
gu_counts = {}
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[2]:
        m = re.search(r'([가-힣]+구)', str(row[2]))
        if m:
            gu = m.group(1)
            gu_counts[gu] = gu_counts.get(gu, 0) + 1

for gu, cnt in sorted(gu_counts.items(), key=lambda x: -x[1]):
    print(f"  {gu}: {cnt}개소")
print(f"\n합계: {sum(gu_counts.values())}개소 / 구 수: {len(gu_counts)}개")
